"""结果输出：单篇 Markdown 批改报告 + 汇总 Excel 成绩表 + 失败清单。"""

from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .grader import GradeResult

HEADER_FILL = PatternFill("solid", fgColor="2F5B7C")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAND_FILL = {
    "优秀": PatternFill("solid", fgColor="D9EAD3"),
    "不及格": PatternFill("solid", fgColor="F4CCCC"),
}


def report_markdown(res: GradeResult, model_note: str = "") -> str:
    lines = [
        f"# 《{res.title}》批改报告",
        "",
        f"- **文件**：{res.file}",
        f"- **论文类型**：{res.rubric_name}",
        f"- **总分**：**{res.total}**（{res.band}）　权重制：{len(res.dimensions)} 个维度加权",
        f"- **评分置信度**：{res.confidence:.2f}" + ("　⚠️ 建议人工复核" if res.confidence < 0.6 else ""),
        f"- **批改方式**：{model_note or ('mock 试运行' if res.mock else 'AI 辅助评分')}",
        f"- **篇幅**：约 {res.n_chars} 字" if res.n_chars else "",
        f"- **标记**：{'；'.join(res.flags) if res.flags else '无'}",
        "",
    ]

    lines += ["## 分维度评分", "",
              "| 维度 | 权重 | 得分 | 加权得分 | 置信度 |",
              "|---|---|---|---|---|"]
    for d in res.dimensions:
        lines.append(
            f"| {d.name} | {d.weight} | {d.score:.0f} | {d.score * d.weight / 100:.1f} | {d.confidence:.2f} |"
        )
    lines.append(f"| **合计** | 100 | — | **{res.total}** | — |")
    lines.append("")

    for d in res.dimensions:
        lines += [f"### {d.name}：{d.score:.0f} 分", "", d.comment or "（无评语）", ""]
        if d.evidence:
            lines += ["**原文证据**："]
            lines += [f"> {e}" for e in d.evidence]
            lines.append("")
        if d.suspicions:
            lines += [f"**⚠️ 学术规范提示**：{d.suspicions}", ""]

    if res.overall_comment:
        lines += ["## 总评", "", res.overall_comment, ""]
    if res.strengths:
        lines += ["## 主要优点"] + [f"{i}. {s}" for i, s in enumerate(res.strengths, 1)] + [""]
    if res.improvements:
        lines += ["## 修改建议"] + [f"{i}. {s}" for i, s in enumerate(res.improvements, 1)] + [""]
    lines += ["---", f"*生成时间：{datetime.now():%Y-%m-%d %H:%M}*"]
    return "\n".join(lines)


def save_report(res: GradeResult, out_dir: Path, model_note: str = "") -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    stem = Path(res.file).stem
    path = out_dir / f"批改报告__{stem}.md"
    path.write_text(report_markdown(res, model_note), encoding="utf-8")
    return path


def _style_header(ws, row_idx: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def save_summary_xlsx(results: list[GradeResult], out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ok = [r for r in results if r.error is None]
    failed = [r for r in results if r.error is not None]

    # ---------- 成绩总表 ----------
    ws = wb.active
    ws.title = "成绩总表"
    cols = ["文件名", "论文题目", "论文类型", "总分", "等级", "置信度", "标记", "备注"]
    ws.append(cols)
    _style_header(ws, 1, len(cols))
    for r in sorted(ok, key=lambda x: -x.total):
        ws.append([
            r.file, r.title, r.rubric_name, r.total, r.band, r.confidence,
            "；".join(r.flags), "mock 试运行" if r.mock else "",
        ])
        row = ws.max_row
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(c in (2, 7)))
        band_cell = ws.cell(row=row, column=5)
        if r.band in BAND_FILL:
            band_cell.fill = BAND_FILL[r.band]
    if failed:
        for r in failed:
            ws.append([r.file, r.title or "", r.rubric_name, "", "", "", "批改失败", r.error])
            for c in range(1, len(cols) + 1):
                ws.cell(row=ws.max_row, column=c).border = BORDER

    widths = [28, 32, 16, 8, 8, 8, 30, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if ok:
        n = len(ok)
        avg = sum(r.total for r in ok) / n
        stats_row = ws.max_row + 2
        bands = {}
        for r in ok:
            bands[r.band] = bands.get(r.band, 0) + 1
        dist = "，".join(f"{b} {n_}人" for b, n_ in bands.items())
        ws.cell(row=stats_row, column=1,
                value=f"统计：共 {n} 篇成功，平均 {avg:.1f} 分，最高 {max(r.total for r in ok)}，"
                      f"最低 {min(r.total for r in ok)}。等级分布：{dist}")
        ws.cell(row=stats_row, column=1).font = Font(bold=True)

    # ---------- 按论文类型分表（维度明细） ----------
    for ptype in {r.ptype for r in ok}:
        group = [r for r in ok if r.ptype == ptype]
        dims = group[0].dimensions
        sheet_name = re.sub(r"[（(].*?[）)]", "", group[0].rubric_name)[:10] + "明细"
        wsr = wb.create_sheet(sheet_name)
        cols = ["文件名", "论文题目", "总分", "等级"] + [
            x for d in dims for x in (f"{d.name}({d.weight})", f"{d.name}评语")]
        wsr.append(cols)
        _style_header(wsr, 1, len(cols))
        for r in sorted(group, key=lambda x: -x.total):
            row_vals = [r.file, r.title, r.total, r.band]
            for d in r.dimensions:
                row_vals += [d.score, d.comment]
            wsr.append(row_vals)
            for c in range(1, len(cols) + 1):
                wsr.cell(row=wsr.max_row, column=c).border = BORDER
        wsr.column_dimensions["A"].width = 28
        wsr.column_dimensions["B"].width = 32
        for c in range(3, len(cols) + 1):
            wsr.column_dimensions[get_column_letter(c)].width = 30
        wsr.freeze_panes = "C2"

    path = out_dir / f"批改成绩汇总_{datetime.now():%Y%m%d_%H%M}.xlsx"
    wb.save(path)
    return path


def save_failures_csv(failed: list[GradeResult], out_dir: Path) -> Path:
    path = out_dir / f"失败清单_{datetime.now():%Y%m%d_%H%M}.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["文件名", "失败原因"])
        for r in failed:
            w.writerow([r.file, r.error])
    return path
